import io
import json
import openpyxl
import tempfile
import os

from typing import Optional, Union
from fastapi.responses import StreamingResponse
from fastapi import UploadFile, File, APIRouter, Depends, Query
from ismcore.messaging.base_message_route_model import RouteMessageStatus
from ismcore.model.base_model import ProcessorStateDirection
from ismcore.model.processor_state import State
from pydantic import ValidationError

from api import token_service
from api.processor_state_route import SELECTOR_STATE_ROUTER
from environment import storage
from message_router import message_router
from utils.http_exceptions import check_null_response
from utils.process_file import process_csv_state_sync_store

state_router = APIRouter()
state_router_route = message_router.find_route(SELECTOR_STATE_ROUTER)

@state_router.get('/{state_id}')
@check_null_response
async def fetch_state(
    state_id: str,
    load_data: bool = False,
    offset: int = Query(..., description="Offset for pagination"),
    limit: int = Query(..., description="Limit for pagination"),
    user_id: str = Depends(token_service.verify_jwt)
) -> Optional[State]:
    return storage.load_state(state_id=state_id, load_data=load_data, offset=offset, limit=limit)


def _write_excel_row(worksheet, row_index: int, row_data: dict, column_mapping: dict):
    """Write a single row to the Excel worksheet."""
    excel_row = row_index + 2  # +2 for 1-based indexing and header row
    for col_name, col_value in row_data.items():
        if col_name in column_mapping:
            worksheet.cell(row=excel_row, column=column_mapping[col_name], value=col_value)


def _process_chunk_rows(rows, worksheet, column_mapping: dict):
    """Process database rows and write them to Excel worksheet."""
    current_row_data = {}
    current_data_index = None

    for col_name, data_index, data_value in rows:
        # New row detected - write previous row to Excel
        if current_data_index is not None and data_index != current_data_index:
            _write_excel_row(worksheet, current_data_index, current_row_data, column_mapping)
            current_row_data = {}

        current_data_index = data_index
        if col_name:
            current_row_data[col_name] = data_value

    # Write the last row in the chunk
    if current_data_index is not None:
        _write_excel_row(worksheet, current_data_index, current_row_data, column_mapping)


def _file_iterator(file_path: str, chunk_size: int = 8192):
    """Generator to read file in chunks and clean up after streaming."""
    try:
        with open(file_path, 'rb') as file:
            while chunk := file.read(chunk_size):
                yield chunk
    finally:
        # Clean up the temporary file after streaming
        if os.path.exists(file_path):
            os.unlink(file_path)

@state_router.get(
    "/{state_id}/export",
    summary="Export state as Excel",
    response_class=StreamingResponse,
)
async def export_state_excel(
    state_id: str,
    chunk_size: int = Query(1000, description="Number of rows to load per chunk"),
    user_id: str = Depends(token_service.verify_jwt)
) -> StreamingResponse:
    # Load state metadata
    state_meta = storage.load_state_metadata(state_id=state_id)
    if not state_meta:
        raise ValueError(f"State {state_id} not found")

    # Initialize Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"State {state_id}"[:31]  # Excel sheet names max 31 chars

    # Write header row
    columns = list(state_meta.columns.keys())
    column_mapping = {col_name: idx + 1 for idx, col_name in enumerate(columns)}
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Fetch and write data in chunks
    offset = 0
    total_count = state_meta.count

    while offset < total_count:
        rows = storage.fetch_state_data_chunk_for_export(
            state_id=state_id,
            offset=offset,
            limit=chunk_size
        )

        if not rows:
            break

        _process_chunk_rows(rows, ws, column_mapping)
        offset += chunk_size

    # Save to temporary file and stream
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp_file:
        tmp_path = tmp_file.name
        wb.save(tmp_file.name)

    return StreamingResponse(
        _file_iterator(tmp_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="state_{state_id}.xlsx"'},
    )

@state_router.post("/create")
@check_null_response
async def merge_state(state: State) -> State:
    # TODO we should propagate this as a CQRS event to the state machine instead of saving it directly
    return storage.save_state(state=state, options={
        "force_update_column": True,
        "force_update_count": False,
    })

@state_router.delete('/{state_id}/data')
@check_null_response
async def delete_state_data(state_id: str) -> int:
    result = storage.delete_state_data(state_id=state_id)
    return 1


@state_router.delete('/{state_id}')
async def delete_state(state_id: str, user_id: str = Depends(token_service.verify_jwt)) -> int:
    storage.delete_processor_state_routes_by_state_id(state_id=state_id)
    storage.delete_state_cascade(state_id=state_id)
    storage.delete_workflow_edges_by_node_id(node_id=state_id)
    storage.delete_workflow_node(node_id=state_id)
    return 1


@state_router.delete("/{state_id}/config/{definition_type}/{id}")
@check_null_response
async def delete_config_definition(state_id: str, definition_type: str, id: str) -> int:
    return storage.delete_state_config_key_definition(
        state_id=state_id,
        definition_type=definition_type,
        definition_id=id
    )


@state_router.delete("/{state_id}/column/{column_id}")
@check_null_response
async def delete_state_column(state_id, column_id) -> int:
    storage.delete_state_column()

    # return storage.save_state(state=state, options={
    #     "force_update_column": True
    # })



def derive_message_from_input_list(route_id: str, query_state: list[str]):
    # for query_state_entry in query_state:
    raise NotImplementedError("input message as list is not implemented")


def derive_message_from_input_dict(route_id: str, query_state_entry: dict):
    response_message = {
        "type": "query_state_entry",
        "route_id": route_id,
        "query_state": [query_state_entry]
    }

    return response_message


def derive_message_from_input_value(route_id: str, query_state_entry_value: any):
    return {
        "type": "query_state_entry",
        "route_id": route_id,
        "query_state": [
            {"input": str(query_state_entry_value)}
        ]
    }

#
# class QueryStateEntry(BaseModel):
#     session_id: Optional[str] = None
#     user: Optional[str] = None
#     input: Optional[str] = None
#     content: Optional[str] = None


@state_router.post('/{state_id}/forward/entry')
@check_null_response
async def route_forward_query_state_entry(state_id: str, input_value: Union[str, dict, bytes]) -> RouteMessageStatus:
    state = storage.fetch_state(state_id=state_id)
    if not state:
        raise ValidationError(f'input state id {state_id} does not exist')

    # fetch the processor state route for the input state id
    # essentially what this is doing is finding a set of processors that take this state as their input
    processor_state_routes = storage.fetch_processor_state_route(
        state_id=state_id,
        direction=ProcessorStateDirection.INPUT)

    if not processor_state_routes:
        raise ValidationError(f"state_id: {state_id} is not connected to any processor inputs")

    # user = "krasaee"  # TODO need to extract from jwt

    status = None # TODO need to fix this such that we do not send hundreds of messages.
    # we iterate each processor state route and submit the input value to the processor
    for processor_state in processor_state_routes:
        route_id = processor_state.id
        if isinstance(input_value, dict):
            message = derive_message_from_input_dict(route_id=route_id, query_state_entry=input_value)
        elif isinstance(input_value, list):
            message = derive_message_from_input_list(route_id=route_id, query_state=input_value)
        else:
            message = derive_message_from_input_value(route_id=input_value)

        # convert to a string object for submission to the pub system
        message_string = json.dumps(message)
        status = await state_router_route.publish(message_string)

    return status


@state_router.post("/{state_id}/data/upload")
async def upload_file(state_id: str, file: UploadFile = File(...)):
    try:
        state = storage.load_state(state_id=state_id)

        if not state:
            raise KeyError(f"unable to locate state id {state_id}")

        data = await file.read()
        text_data = io.StringIO(data.decode('utf-8'))
        query_state = await process_csv_state_sync_store(state=state, io=text_data)

        ## publish blocks of data instead of a one shot dataset
        offset = 0
        chunks = 200
        sync_route = message_router.find_route("processor/state/sync")
        while offset < len(query_state):
            block = query_state[offset:offset + chunks]
            offset += chunks

            # derive the new message with complete csv file data
            message = {
                "type": "query_state_direct",
                "state_id": state.id,
                "query_state": block
            }

            message_string = json.dumps(message)
            await sync_route.publish(msg=message_string)


        # sync_route.flush()

        # state = storage.save_state(state=state)

        return {
            "status": "success",
            "message": "file uploaded successfully",
            "state_id": state.id,
            "count": state.count
        }
    except Exception as e:
        return {"status": "error", "message": str(e), "count": 0}


@state_router.get('/{state_id}/processors')
@check_null_response
async def fetch(state_id: str):
    return storage.fetch_processor_state(state_id=state_id)
